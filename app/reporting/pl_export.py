"""Exports a live, formula-driven P&L workbook -- a "GL Data" tab holding
the entity's full ledger, and a "P&L" tab where every figure (line items,
subtotals, gross profit, operating income, net income) is a real Excel
formula referencing that data, not a pasted number. Change the period
cells or the underlying GL Data and the whole report recalculates.

This is the honest substitute for a proprietary tool's live-data Excel
functions (e.g. DataRails' `DR.GET`): SUMIFS is a standard, verifiable
Excel function that achieves the same "live, not hardcoded" outcome
without guessing at syntax we can't confirm. If a client has a real
DataRails license, this GL Data tab is exactly the kind of source table
their own functions would point at.
"""
from __future__ import annotations

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.coa import PL_ACCOUNT_TYPES, AccountType, CREDIT_NORMAL_TYPES, ChartOfAccounts
from app.models import GLEntry

GL_DATA_MAX_ROW = 100_000  # bounded range for SUMIFS -- generous headroom, not a full-column scan

_SECTION_TITLES = {
    AccountType.REVENUE: "REVENUE",
    AccountType.COGS: "COST OF GOODS SOLD",
    AccountType.OPERATING_EXPENSE: "OPERATING EXPENSES",
    AccountType.OTHER_INCOME: "OTHER INCOME",
    AccountType.OTHER_EXPENSE: "OTHER EXPENSE",
}

_BOLD = Font(bold=True)
_HEADER_FONT = Font(bold=True, size=14)


def _write_gl_data_sheet(ws: Worksheet, gl_entries: list[GLEntry]) -> None:
    ws.append(["Date", "Account Code", "Account Name", "Amount"])
    for cell in ws[1]:
        cell.font = _BOLD
    for entry in sorted(gl_entries, key=lambda e: (e.date, e.account_code)):
        ws.append([entry.date, entry.account_code, entry.account_name, float(entry.amount)])
    for column_cells in ws.columns:
        length = max(len(str(c.value)) for c in column_cells if c.value is not None)
        ws.column_dimensions[column_cells[0].column_letter].width = max(12, length + 2)


def _sumifs_formula(account_code: str, negate: bool) -> str:
    date_col = f"'GL Data'!$A$2:$A${GL_DATA_MAX_ROW}"
    account_col = f"'GL Data'!$B$2:$B${GL_DATA_MAX_ROW}"
    amount_col = f"'GL Data'!$D$2:$D${GL_DATA_MAX_ROW}"
    inner = (
        f'SUMIFS({amount_col}, {account_col}, "{account_code}", '
        f'{date_col}, ">="&$B$2, {date_col}, "<="&$B$3)'
    )
    return f"=-{inner}" if negate else f"={inner}"


def _write_pl_sheet(
    ws: Worksheet,
    chart_of_accounts: ChartOfAccounts,
    entity_id: str,
    entity_name: str,
    period_start: date,
    period_end: date,
    unclassified_account_codes: list[str],
) -> None:
    ws["A1"] = f"Profit & Loss -- {entity_name}"
    ws["A1"].font = _HEADER_FONT
    ws["A2"] = "Period Start"
    ws["B2"] = period_start
    ws["A3"] = "Period End"
    ws["B3"] = period_end
    for ref in ("A1", "A2", "A3"):
        ws[ref].font = _BOLD

    accounts_by_type: dict[AccountType, list] = {t: [] for t in PL_ACCOUNT_TYPES}
    for entry in sorted(chart_of_accounts.accounts_for(entity_id), key=lambda e: e.account_code):
        if entry.account_type in PL_ACCOUNT_TYPES:
            accounts_by_type[entry.account_type].append(entry)

    row = 5
    subtotal_rows: dict[AccountType, int] = {}

    def _write_section(account_type: AccountType) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=_SECTION_TITLES[account_type]).font = _BOLD
        row += 1
        first_data_row = row
        for entry in accounts_by_type[account_type]:
            ws.cell(row=row, column=1, value=f"{entry.account_name} ({entry.account_code})")
            formula = _sumifs_formula(entry.account_code, negate=account_type in CREDIT_NORMAL_TYPES)
            ws.cell(row=row, column=2, value=formula)
            row += 1
        total_label_row = row
        ws.cell(row=total_label_row, column=1, value=f"Total {_SECTION_TITLES[account_type].title()}").font = _BOLD
        if row > first_data_row:
            ws.cell(row=total_label_row, column=2, value=f"=SUM(B{first_data_row}:B{row - 1})").font = _BOLD
        else:
            ws.cell(row=total_label_row, column=2, value=0).font = _BOLD
        subtotal_rows[account_type] = total_label_row
        row += 2

    _write_section(AccountType.REVENUE)
    _write_section(AccountType.COGS)

    gross_profit_row = row
    ws.cell(row=row, column=1, value="Gross Profit").font = _BOLD
    ws.cell(row=row, column=2, value=f"=B{subtotal_rows[AccountType.REVENUE]}-B{subtotal_rows[AccountType.COGS]}").font = _BOLD
    row += 2

    _write_section(AccountType.OPERATING_EXPENSE)

    operating_income_row = row
    ws.cell(row=row, column=1, value="Operating Income").font = _BOLD
    ws.cell(
        row=row, column=2,
        value=f"=B{gross_profit_row}-B{subtotal_rows[AccountType.OPERATING_EXPENSE]}",
    ).font = _BOLD
    row += 2

    _write_section(AccountType.OTHER_INCOME)
    _write_section(AccountType.OTHER_EXPENSE)

    net_income_row = row
    ws.cell(row=row, column=1, value="NET INCOME").font = _BOLD
    ws.cell(
        row=row, column=2,
        value=(
            f"=B{operating_income_row}+B{subtotal_rows[AccountType.OTHER_INCOME]}"
            f"-B{subtotal_rows[AccountType.OTHER_EXPENSE]}"
        ),
    ).font = _BOLD
    row += 2

    if unclassified_account_codes:
        ws.cell(row=row, column=1, value="Unclassified activity (excluded above -- classify in chart of accounts)").font = _BOLD
        row += 1
        for code in unclassified_account_codes:
            ws.cell(row=row, column=1, value=code)
            ws.cell(row=row, column=2, value=_sumifs_formula(code, negate=False))
            row += 1

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 18
    ws["B2"].number_format = "yyyy-mm-dd"
    ws["B3"].number_format = "yyyy-mm-dd"
    for r in range(5, row):
        ws.cell(row=r, column=2).number_format = "#,##0.00"


def build_pl_workbook(
    gl_entries: list[GLEntry],
    chart_of_accounts: ChartOfAccounts,
    entity_id: str,
    entity_name: str,
    period_start: date,
    period_end: date,
) -> Workbook:
    classified_codes = {a.account_code for a in chart_of_accounts.accounts_for(entity_id)}
    unclassified_account_codes = sorted({e.account_code for e in gl_entries if e.account_code not in classified_codes})

    wb = Workbook()
    gl_ws = wb.active
    gl_ws.title = "GL Data"
    _write_gl_data_sheet(gl_ws, gl_entries)

    pl_ws = wb.create_sheet("P&L")
    _write_pl_sheet(
        pl_ws, chart_of_accounts, entity_id, entity_name, period_start, period_end, unclassified_account_codes
    )
    wb.active = wb.sheetnames.index("P&L")

    return wb
